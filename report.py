import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from collections import defaultdict
from datetime import datetime
import os

def calculate_attendance_percentage(file_names):
    attendance_count = defaultdict(lambda: [0, 0])  # Default values are [present days, total days]

    for file_name in file_names:
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

            # Scan unique dates to calculate total working days
            unique_dates = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                date_str = row[4]
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                unique_dates.add(date)

            # Count the present days and total days for each student
            for row in sheet.iter_rows(min_row=2, values_only=True):
                roll_number = row[0]
                student_name = row[1]
                attendance = row[2]

                attendance_count[(roll_number, student_name)][1] += 1  # Increment total days

                if attendance == 'Present':
                    attendance_count[(roll_number, student_name)][0] += 1  # Increment present days

        except Exception as e:
            print("Error reading Excel file:", e)

    # Calculate attendance percentage for each student
    attendance_percentage = {}
    for (roll_number, student_name), (present_days, total_days) in attendance_count.items():
        percentage = (present_days / total_days) * 100
        attendance_percentage[(roll_number, student_name)] = round(percentage, 2)

    return attendance_percentage

def create_attendance_report(attendance_percentage, file_name):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add headings
        headings = ['Roll Number', 'Name', 'Attendance Percentage']
        for col_num, heading in enumerate(headings, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet.cell(row=1, column=col_num)
            cell.value = heading
            cell.font = Font(bold=True)

        # Write roll number, student name, and attendance percentage for each student
        for row_num, ((roll_number, student_name), percentage) in enumerate(attendance_percentage.items(), 2):
            sheet.cell(row=row_num, column=1, value=roll_number)
            sheet.cell(row=row_num, column=2, value=student_name)
            sheet.cell(row=row_num, column=3, value=percentage)

        # Adjust column widths
        for col_num in range(1, len(headings) + 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 20

        # Create the "attendance_report" folder if it doesn't exist
        if not os.path.exists("Attendance_Report"):
            os.makedirs("Attendance_Report")

        # Save the report in the "attendance_report" folder
        report_file_name = os.path.join("Attendance_Report", os.path.splitext(os.path.basename(file_name))[0] + '_report.xlsx')
        workbook.save(report_file_name)
        print(f"Attendance report saved as '{report_file_name}'")

    except Exception as e:
        print("Error creating attendance report:", e)

# Usage example
file_names = ['Attendance_Files/Btech_CS_4th/KHU-801.xlsx', 'Attendance_Files/Btech_CS_4th/KOE-083.xlsx', 'Attendance_Files/Btech_CS_4th/KOE-094.xlsx']

for file_name in file_names:
    attendance_percentage = calculate_attendance_percentage([file_name])
    create_attendance_report(attendance_percentage, file_name)
    print(f"Attendance report created for file: {file_name}")
    print()
