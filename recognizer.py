import cv2
import os
import numpy as np
from datetime import datetime, date, timedelta
import openpyxl
import sqlite3

# Establish a connection to the SQLite database
db_connection = sqlite3.connect('database/database.db')

def fetch_student_details(id):
    # Fetch student details from the database
    try:
        # Create a cursor to execute SQL queries
        cursor = db_connection.cursor()

        # Fetch the student details based on the Roll Number
        query = "SELECT Name, Roll_number FROM students WHERE Roll_number = ?"
        cursor.execute(query, (id,))
        result = cursor.fetchone()

        # Close the cursor
        cursor.close()

        return result
    except sqlite3.Error as e:
        print("Error fetching student details:", e)

    return None

def fetch_all_student_ids():
    # Fetch all student IDs from the database
    try:
        # Create a cursor to execute SQL queries
        cursor = db_connection.cursor()

        # Fetch all the student IDs
        query = "SELECT Roll_number FROM students"
        cursor.execute(query)
        result = cursor.fetchall()

        # Close the cursor
        cursor.close()

        # Extract the IDs from the result
        ids = [row[0] for row in result]

        return ids
    except sqlite3.Error as e:
        print("Error fetching student IDs:", e)

    return []

def write_details_to_excel(recognized_ids, file_name, attendance_date):
    # Write attendance details to the Excel file
    try:
        # Check if the Excel file exists
        if not os.path.isfile(file_name):
            # Create a new Excel file with headings if it doesn't exist
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Set the headings in the first row
            headings = ['Roll Number', 'Name', 'Attendance', 'Time', 'Date']
            sheet.append(headings)
        else:
            # Load the existing Excel file
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

        # Fetch the student details for each recognized ID
        for id in recognized_ids:
            student_details = fetch_student_details(id)

            if student_details:
                name = student_details[0]
                rollnumber = student_details[1]

                # Get the current time
                current_time = datetime.now().time().strftime('%H:%M')

                # Check if the ID has already been marked for the given date
                is_marked = False
                for row in sheet.iter_rows(min_row=2, max_col=5, max_row=sheet.max_row):
                    if row[0].value == rollnumber and row[4].value == attendance_date.strftime('%Y-%m-%d'):
                        is_marked = True
                        break

                if not is_marked:  # Only mark attendance if not already marked
                    # Write the details to the Excel sheet
                    last_row = sheet.max_row + 1  # Get the next available row
                    sheet.cell(row=last_row, column=1, value=rollnumber)
                    sheet.cell(row=last_row, column=2, value=name[:20])  # Limit name to 20 characters
                    sheet.cell(row=last_row, column=3, value="Present"[:12])  # Limit attendance to 12 characters
                    sheet.cell(row=last_row, column=4, value=current_time)
                    sheet.cell(row=last_row, column=5, value=attendance_date.strftime('%Y-%m-%d'))

                    # Set the date format and column widths
                    date_cell = sheet.cell(row=last_row, column=5)
                    date_cell.number_format = 'yyyy-mm-dd'
                    sheet.column_dimensions['E'].width = 15
                    sheet.column_dimensions['A'].width = 15  # Set the width of roll number column
                    sheet.column_dimensions['B'].width = 20  # Set the width of name column
                    sheet.column_dimensions['C'].width = 12  # Set the width of attendance column

        # Save the changes to the Excel file
        workbook.save(file_name)
    except Exception as e:
        print("Error writing details to Excel:", e)


def mark_absent_students(recognized_ids, all_ids, file_name, attendance_date):
    # Mark absent students in the Excel file
    try:
        # Load the existing or newly created Excel file
        if not os.path.isfile(file_name):
            # Create a new Excel file with headings if it doesn't exist
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Set the headings in the first row
            headings = ['Roll Number', 'Name', 'Attendance', 'Time', 'Date']
            sheet.append(headings)
        else:
            # Load the existing Excel file
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

        # Fetch the student details for each absent ID
        for id in all_ids:
            if id not in recognized_ids:
                student_details = fetch_student_details(id)

                if student_details:
                    name = student_details[0]
                    rollnumber = student_details[1]

                    # Check if the ID has already been marked for the given date
                    is_marked = False
                    for row in sheet.iter_rows(min_row=2, max_col=5, max_row=sheet.max_row):
                        if row[0].value == rollnumber and row[4].value == attendance_date.strftime('%Y-%m-%d'):
                            is_marked = True
                            break

                    if not is_marked:  # Only mark absent if not already marked
                        # Write the details to the Excel sheet
                        last_row = sheet.max_row + 1  # Get the next available row
                        sheet.cell(row=last_row, column=1, value=rollnumber)
                        sheet.cell(row=last_row, column=2, value=name[:20])  # Limit name to 20 characters
                        sheet.cell(row=last_row, column=3, value="Absent"[:12])  # Limit attendance to 12 characters
                        sheet.cell(row=last_row, column=4, value='')
                        sheet.cell(row=last_row, column=5, value=attendance_date.strftime('%Y-%m-%d'))

                        # Set the date format and column widths
                        date_cell = sheet.cell(row=last_row, column=5)
                        date_cell.number_format = 'yyyy-mm-dd'
                        sheet.column_dimensions['E'].width = 15
                        sheet.column_dimensions['A'].width = 15  # Set the width of roll number column
                        sheet.column_dimensions['B'].width = 20  # Set the width of name column
                        sheet.column_dimensions['C'].width = 12  # Set the width of attendance column

        # Save the changes to the Excel file
        workbook.save(file_name)
    except Exception as e:
        print("Error marking absent students in Excel:", e)
        

# Instance for cascade file
face_cas = cv2.CascadeClassifier('cascades/haarcascade_frontalface_default.xml')

# Initializing camera for the time until it records 30 images into the dataset
cap_video = cv2.VideoCapture(0)

# Creating an instance for LBPH algorithm
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Reading trainer.yml file
recognizer.read('trainer/trainer.yml')

# Selecting font
font = cv2.FONT_HERSHEY_SIMPLEX

# Set the time ranges for face recognition
time_ranges = [
    {
        'start': datetime.strptime('09:00 AM', '%I:%M %p').time(),
        'end': datetime.strptime('09:50 AM', '%I:%M %p').time(),
        'file': 'Attendance_Files/Btech_CS_4th/KHU-801.xlsx'
    },
    {
        'start': datetime.strptime('09:51 AM', '%I:%M %p').time(),
        'end': datetime.strptime('10:40 AM', '%I:%M %p').time(),
        'file': 'Attendance_Files/Btech_CS_4th/KOE-083.xlsx'
    },
    {
        'start': datetime.strptime('11:00 AM', '%I:%M %p').time(),
        'end': datetime.strptime('03:00 PM', '%I:%M %p').time(),
        'file': 'Attendance_Files/Btech_CS_4th/KOE-094.xlsx'
    }
]
# Fetch all student IDs from the database
all_ids = fetch_all_student_ids()
# Get the current time
current_time = datetime.now().time()

# Check if the current time falls within any of the defined time ranges
for time_range in time_ranges:
    recognized_ids = []
    start_time = time_range['start']
    end_time = time_range['end']
    file_name = time_range['file']

    if start_time <= current_time <= end_time:
        # Face recognition loop for the current time range
        while True:
            # Capture video frame
            ret, img = cap_video.read()

            if not ret:
                print("Failed to capture video frame")
                break

            # Convert color BGR to grayscale format
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # Detect faces in the grayscale image
            faces = face_cas.detectMultiScale(gray, 1.3, 7)

            for (x, y, w, h) in faces:
                # Extract the face region of interest (ROI)
                roi_gray = gray[y:y + h, x:x + w]

                # Recognize the face using the LBPH recognizer
                id, _ = recognizer.predict(roi_gray)

                if id not in recognized_ids:
                    recognized_ids.append(id)  # Add recognized ID to the array

                    # Get the current date
                    attendance_date = datetime.now().date()

                    write_details_to_excel(recognized_ids, file_name, attendance_date)  # Write details to the Excel file

                    # Draw a rectangle around the face and display the ID
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    cv2.putText(img, str(id), (x, y - 10), font, 0.55, (120, 255, 120), 1)

            # Display the frame with face detections
            cv2.imshow('Face Recognition', img)
            # Get the current date
            attendance_date = datetime.now().date()

            # Check if the current time is greater than the end time
            if datetime.now().time() >= end_time:
                # Write details to the Excel file
                write_details_to_excel(recognized_ids, file_name, attendance_date)

                # Mark absent students in the Excel file
                mark_absent_students(recognized_ids, all_ids, file_name, attendance_date)
                break
            # Break the loop when 'q' is pressed
            if cv2.waitKey(1) & 0xFF == ord('q'):
                # Write details to the Excel file
                write_details_to_excel(recognized_ids, file_name, attendance_date)

                # Mark absent students in the Excel file
                mark_absent_students(recognized_ids, all_ids, file_name, attendance_date)
                break
        break

# Close the database connection
db_connection.close()

# Release the video capture and destroy all windows
cap_video.release()
cv2.destroyAllWindows()